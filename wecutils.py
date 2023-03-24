import os
import datetime
import copy
from pathlib import Path
from openpyxl import load_workbook

from Event import Event
from Year import Year
from Month import Month


# ---------------------------- Math ------------------------------ #
def total_all(data):
    """Returns a dictionary containing each event with a corresponding running total computed from the
    given Excel sheets."""
    total_all_dict = {}
    for item in data:
        if total_all_dict.get(item.name) is not None:
            total_all_dict[item.name] += 1
        else:
            total_all_dict[item.name] = 1
    return dict(sorted(total_all_dict.items(), key=lambda x: x[1], reverse=True))


def total_byyear(data):
    """Returns a dictionary containing each year represented in the data from the given Excel sheets. Each year
    is mapped to a dictionary containing totals for each event from that year."""
    total_byyear_dict = {}
    temp = {}
    for year in data:
        if total_byyear_dict.get(year.year) is None:
            total_byyear_dict[year.year] = {}
        for item in year.lst:
            if total_byyear_dict[year.year].get(item.name) is not None:
                total_byyear_dict[year.year][item.name] += 1
            else:
                total_byyear_dict[year.year][item.name] = 1
        # sort each year in dict into new dict
        temp[year.year] = dict(sorted(total_byyear_dict[year.year].items(), reverse=True, key=lambda x: x[1]))
    return temp


def total_bymonth(data):
    """Returns a dictionary containing each month represented in the data from the given Excel sheets. Each month
    is mapped to a dictionary containing totals for each event from that month."""
    total_bymonth_dict = {}
    temp = {}
    for month in data:
        if total_bymonth_dict.get(month.month) is None:
            total_bymonth_dict[month.month] = {}
        for event in month.lst:
            if total_bymonth_dict[month.month].get(event.name) is not None:
                total_bymonth_dict[month.month][event.name] += 1
            else:
                total_bymonth_dict[month.month][event.name] = 1
        temp[month.month] = dict(sorted(total_bymonth_dict[month.month].items(), reverse=True, key=lambda x: x[1]))
    return temp


def avg_bymonth(data_y, tbm):
    """Computes the average number of events in each month and returns the results as a dictionary."""
    avg_bymonth_dict = copy.deepcopy(tbm)
    months_in_last_year = [event.month for event in data_y[-1].lst]
    for month in avg_bymonth_dict:
        for event in avg_bymonth_dict[month]:
            if month in months_in_last_year:
                avg_bymonth_dict[month][event] = round(avg_bymonth_dict[month][event] / len(data_y), 5)
            else:
                avg_bymonth_dict[month][event] = round(avg_bymonth_dict[month][event] / (len(data_y) - 1), 5)
    return avg_bymonth_dict


def percent_total(ta):
    """Computes a sum total of all events and computes the percentage of each event against that total."""
    total_events = 0
    percent_total_dict = {}
    for event in ta:
        total_events += ta[event]
    for event in ta:
        percent_total_dict[event] = (ta[event] / total_events) * 100
    return percent_total_dict


def percent_bytime(tbt):
    """Computes a sum total of all events in each year and computes the percentage of each event against that total."""
    percent_bytime_dict = {}
    for per in tbt:
        total_events = 0
        percent_bytime_dict[per] = {}
        for item in tbt[per]:
            total_events += tbt[per][item]
        for item in tbt[per]:
            percent_bytime_dict[per][item] = (tbt[per][item] / total_events) * 100
    return percent_bytime_dict
# ---------------------------------------------------------------- #


# ----------------------- Event Handling ------------------------- #
def map_to_alias(event_name, aliases):
    """Maps a given event name to a corresponding name within the 'aliases.txt' file"""
    if aliases.get(event_name) is not None:
        mapped_name = aliases[event_name]
        return mapped_name.upper()
    else:
        return event_name.upper()


def read_alias_file():
    """Reads the aliases.txt file and parses out each given alias. Returns a dictionary."""
    aliases = {}
    curdir = os.getcwd()
    os.chdir(Path(curdir).parent)
    f = open("aliases.txt", "r")
    for line in f:
        key_value = line.upper().strip().split('=')
        aliases[key_value[0]] = key_value[1]
    os.chdir(curdir)
    return aliases
# ---------------------------------------------------------------- #


# ------------------------ Error Logging ------------------------- #
temp_error_log = []


def log_error(e, year, cal_name):
    """Adds the given error to the temp_error_log list"""
    temp_error_log.append(f"{str(datetime.datetime.now())} | Error in Calendar '{cal_name}': "
                          f"Failed to add event '{e[0]}' in month '{e[1]}' ({map_to_name(e[1])}) in year '{year}' \n")


def log_name(now):
    """Returns string representation of today's date in the format: mmddyyyy-time"""
    day = str(now.day)
    month = str(now.month)
    year = str(now.year)
    time = str(now.hour) + str(now.minute) + str(now.second) + str(now.microsecond)
    return month + day + year + "-" + time


def write_log():
    """Prints temp_error_list to a .txt file with a name defined by log_name"""
    currf = os.getcwd()
    os.chdir(str(Path(os.getcwd()).parent) + "\\logs")
    f = open(f"{log_name(datetime.datetime.now())}.txt", "a")
    for e in temp_error_log:
        f.write(e)
    f.close()
    os.chdir(currf)
# ---------------------------------------------------------------- #


# --------------------- Reading excel sheets --------------------- #
def parse_month(val):
    """Method for parsing the month out of the given column from the Excel sheet."""
    month = val.split(" ")[0].split("/")[1]
    return month


def parse_year(val):
    """Method for parsing the year out of the given column from the Excel sheet."""
    year = val.split(" ")[0].split("/")[0]
    return int(year)


def read_xlsx(file_lst, aliases):
    """Reads every .xlsx file found in the 'data' directory, separates the year, month, and event name,
    and adds each to a tuple. This tuple is then added to a list and passed to parse_data."""
    line_lst = []
    for file in file_lst:
        sheet = load_workbook(file).active
        for row in sheet.iter_rows(min_row=2, min_col=2, max_col=7):
            if row[5].value is not None:  # skips any rows where the event name cell is empty
                year = parse_year(row[0].value)
                month = parse_month(row[0].value)
                event = row[5].value
                row_tup = (year, month, map_to_alias(event.upper(), aliases))
                line_lst.append(row_tup)
    return parse_data(sorted(line_lst, key=lambda x: (x[0], x[1])), aliases)


def parse_data(data, aliases):
    """Using data list from read_xlsx, creates separate event lists for overall, by year, and by month, returning
    each list in a tuple."""
    master_list = []
    years = []
    months = []
    ###
    for line in data:
        ###
        year = line[0]  # gets year from cell containing date of report
        month = line[1]  # gets month from cell containing date of report
        event = Event(year, month, map_to_alias(line[2].upper(), aliases))  # new event object with year, month, and event name from cell containing event name
        master_list.append(event)  # add new event to master list
        ###
        if year not in [y.year for y in years]:  # create list of years represented in data
            years.append(Year(year))
        if month not in [m.month for m in months]:
            months.append(Month(month))
    ###
    for y in years:
        # add event to each year when event.month is equal to month
        y.add([event for event in master_list if event.year == y.year])
    for m in months:
        # add event to each month when event.month is equal to month
        m.add([event for event in master_list if event.month == m.month])
    ###
    return master_list, years, sorted(months, key=lambda mon: int(mon.month))  # master_list, years, and months are returned to event_counter
# ---------------------------------------------------------------- #


# ---------------------------- Misc ------------------------------ #
def map_to_name(month_num):
    """Maps a given month number code to that month's name."""
    name_dict = {
        "01": "JANUARY",
        "02": "FEBRUARY",
        "03": "MARCH",
        "04": "APRIL",
        "05": "MAY",
        "06": "JUNE",
        "07": "JULY",
        "08": "AUGUST",
        "09": "SEPTEMBER",
        "10": "OCTOBER",
        "11": "NOVEMBER",
        "12": "DECEMBER"
    }
    return name_dict[month_num]
# ---------------------------------------------------------------- #
