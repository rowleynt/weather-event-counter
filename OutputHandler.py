from wecutils import total_all, total_byyear, total_bymonth, avg_bymonth, percent_total, percent_bytime, map_to_name
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Alignment
import os
import datetime


class OutputHandler(object):
    def __init__(self, data):
        self.total_all = total_all(data[0])
        self.total_byyear = total_byyear(data[1])
        self.total_bymonth = total_bymonth(data[2])
        self.avg_bymonth = avg_bymonth(data[1], self.total_bymonth)
        self.percent_total = percent_total(self.total_all)
        self.percent_byyear = percent_bytime(self.total_byyear)
        self.percent_bymonth = percent_bytime(self.total_bymonth)

        self.output_filename = self.output_xlsx()

    def output_xlsx(self):
        filename = self.make_filename(datetime.datetime.now())
        curdir = os.getcwd()
        os.chdir(Path(curdir).parent)
        os.chdir("output")
        wb = Workbook()
        sheet = wb.active
        sheet.title = "output-" + filename
        self.write_to_sheet(sheet)
        wb.save(f"output-{filename}.xlsx")
        os.chdir(curdir)
        return filename

    def write_to_sheet(self, sheet):
        self.format_titles(sheet)
        i = 2
        j = 1
        k = ord('A')
        # ----- writing events from total_byyear ----- #
        for year in self.total_byyear:
            sheet.cell(row=i, column=j).value = year
            sheet.merge_cells(f'{chr(k)}{i}:{chr(k + 1)}{i}')
            sheet.cell(row=i, column=j).alignment = Alignment(horizontal='center')
            i += 1
            for event in self.total_byyear[year]:
                sheet.cell(row=i, column=j).value = event
                sheet.cell(row=i, column=j + 1).value = self.total_byyear[year][event]
                sheet.cell(row=i, column=j + 1).alignment = Alignment(horizontal='left')
                i += 1
        j += 2
        k += 2
        i = 2  # reset i
        # ---- writing events from total_bymonth ----- #
        for month in self.total_bymonth:
            sheet.cell(row=i, column=j).value = map_to_name(month)
            sheet.merge_cells(f'{chr(k)}{i}:{chr(k + 1)}{i}')
            sheet.cell(row=i, column=j).alignment = Alignment(horizontal='center')
            i += 1
            for event in self.total_bymonth[month]:
                sheet.cell(row=i, column=j).value = event
                sheet.cell(row=i, column=j + 1).value = self.total_bymonth[month][event]
                sheet.cell(row=i, column=j + 1).alignment = Alignment(horizontal='left')
                i += 1
        j += 2
        k += 2
        i = 2
        # ------ writing events from total_all ------- #
        sheet.merge_cells(f'{chr(k)}{i}:{chr(k + 1)}{i}')
        i += 1
        for event in self.total_all:
            sheet.cell(row=i, column=j).value = event
            sheet.cell(row=i, column=j + 1).value = self.total_all[event]
            sheet.cell(row=i, column=j + 1).alignment = Alignment(horizontal='left')
            i += 1
        j += 2
        k += 2
        i = 2
        # ----- writing events from avg_bymonth ------ #
        for month in self.avg_bymonth:
            sheet.cell(row=i, column=j).value = map_to_name(month)
            sheet.merge_cells(f'{chr(k)}{i}:{chr(k + 1)}{i}')
            sheet.cell(row=i, column=j).alignment = Alignment(horizontal='center')
            i += 1
            for event in self.avg_bymonth[month]:
                sheet.cell(row=i, column=j).value = event
                sheet.cell(row=i, column=j + 1).value = round(self.avg_bymonth[month][event], 3)
                sheet.cell(row=i, column=j + 1).alignment = Alignment(horizontal='left')
                i += 1
        j += 2
        k += 2
        i = 2
        # ---- writing events from percent_byyear ---- #
        for year in self.percent_byyear:
            sheet.cell(row=i, column=j).value = year
            sheet.merge_cells(f'{chr(k)}{i}:{chr(k + 1)}{i}')
            sheet.cell(row=i, column=j).alignment = Alignment(horizontal='center')
            i += 1
            for event in self.percent_byyear[year]:
                sheet.cell(row=i, column=j).value = event
                sheet.cell(row=i, column=j + 1).value = f'{round(self.percent_byyear[year][event], 5)}%'
                sheet.cell(row=i, column=j + 1).alignment = Alignment(horizontal='left')
                i += 1
        j += 2
        k += 2
        i = 2
        # --- writing events from percent_bymonth ---- #
        for month in self.percent_bymonth:
            sheet.cell(row=i, column=j).value = map_to_name(month)
            sheet.merge_cells(f'{chr(k)}{i}:{chr(k + 1)}{i}')
            sheet.cell(row=i, column=j).alignment = Alignment(horizontal='center')
            i += 1
            for event in self.percent_bymonth[month]:
                sheet.cell(row=i, column=j).value = event
                sheet.cell(row=i, column=j + 1).value = f'{round(self.percent_bymonth[month][event], 5)}%'
                sheet.cell(row=i, column=j + 1).alignment = Alignment(horizontal='left')
                i += 1
        j += 2
        k += 2
        i = 2
        # ---- writing events from percent_total ----- #
        sheet.merge_cells(f'{chr(k)}{i}:{chr(k + 1)}{i}')
        i += 1
        for event in self.percent_total:
            sheet.cell(row=i, column=j).value = event
            sheet.cell(row=i, column=j + 1).value = f'{round(self.percent_total[event], 5)}%'
            sheet.cell(row=i, column=j + 1).alignment = Alignment(horizontal='left')
            i += 1

    @staticmethod
    def format_titles(sheet):
        titles = [
                  "Total Events By Year",
                  "Total Events By Month",
                  "Total Events Overall",
                  "Average Events By Month",
                  "Percentage of Events By Year",
                  "Percentage of Events By Month",
                  "Percentage of Events Overall"
                  ]
        k = 1
        for j in range(len(titles)):
            sheet.cell(row=1, column=k).value = titles[j]
            sheet.cell(row=1, column=k).alignment = Alignment(horizontal='center')
            sheet.cell(row=1, column=k + 1).alignment = Alignment(horizontal='center')
            k += 2
        for i in range(ord('A'), ord('A') + 14, 2):
            sheet.column_dimensions[chr(i)].width = 22
            sheet.column_dimensions[chr(i + 1)].width = 22
            sheet.merge_cells(f'{chr(i)}1:{chr(i + 1)}1')

    @staticmethod
    def make_filename(now):
        day = str(now.day)
        month = str(now.month)
        year = str(now.year)
        time = str(now.hour) + str(now.minute) + str(now.second) + str(now.microsecond)
        return month + day + year + "-" + time

    def __str__(self):
        return f"Successfully generated output file 'output-{self.output_filename}.xlsx' in the 'output' folder."
